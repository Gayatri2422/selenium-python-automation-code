import softest
import logging
import inspect
from openpyxl import load_workbook
import csv

class Utils(softest.TestCase):
     
    def assertListItemText(self,list,value): 
        for stop in list: 
            print("the text is" + stop.text) 
            self.soft_assert(self.assertEqual, stop.text, value)
            if stop.text == value:
                print("test pass")
            else:
                print("test fail")
        self.assert_all()


    def custom_logger(loglevel=logging.DEBUG):
        
        # Set class /method from where its called
        logger_name = inspect.stack()[1][3]
        # create a logger 
        logger = logging.getLogger(logger_name)
        #create a file handler and set the log level
        logger.setLevel(loglevel)
        FH = logging.FileHandler("flight_booking_automation.log")
        # create formatter - how you want to show your logs
        formatter1 = logging.Formatter("%(asctime)s - %(levelname)s - %(name)s :%(filename)s : %(message)s",datefmt="%m/%d/%Y %I:%M:%S %p")        
        # add formatter to console or file handler 
        FH.setFormatter(formatter1)
        # add console handler to logger
        logger.addHandler(FH)

        return logger   
    
    def read_data_from_excel(file_name,sheet):
        datalist = []
        wb = load_workbook(filename=file_name)
        sh = wb[sheet]

        row_ct = sh.max_row
        column_ct = sh.max_column

        for i in range(2,row_ct+1):
            row = []
            for j in range(1,column_ct+1):
                row.append(sh.cell(row=i ,column=j).value)
            datalist.append(row)
        return datalist
    

    def read_data_from_csv(filename):
        #create a empty list
        datalist = []

        #open csv file
        csvdata = open(filename,"r")

        #create a csv_reader
        reader = csv.reader(csvdata)

        # skip the haeder
        next(reader)

        # add csv_row to list
        for rows in reader:
            datalist.append(rows)
        
        return datalist

            
